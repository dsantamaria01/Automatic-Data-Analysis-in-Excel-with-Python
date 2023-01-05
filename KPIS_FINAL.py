#!/usr/bin/env python
# coding: utf-8

# In[54]:


import pandas as pd
from pandas import Series, DataFrame
import fnmatch
import datetime 
import os 
import xlrd
from xlutils.copy import copy
from openpyxl import load_workbook
import pandas as pd
from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains 
from selenium.webdriver.common.keys import Keys # para el enter
import shutil


site = input('introduzca el numero de site,   :')
# Destination path
#r'C:\Users\davalvar\OneDrive - Nokia\Snapshots\SNAPSHOTS 2023\01 Janeiro\05\2900240\2900240_KPIS.xlsx'
path1 =  r'C:\Users\davalvar\OneDrive - Nokia\Snapshots\SNAPSHOTS 2023' 


dia,mes = fecha()

if len(str(site))==7:
    destination=os.path.join(path1,str(mes),str(dia),str(site),str(site)+'__KPIS.xlsx')
else:
    sitep = '0'+str(site)
    destination=os.path.join(path1,str(mes),str(dia),str(sitep),str(sitep)+'__KPIS.xlsx')
    
    
path1 = r'C:\Users\davalvar\Desktop'
path2=os.path.join(path1,str(site))
# Source path
source = r"C:\Users\davalvar\Desktop\Explicaciones Nokia\EXCEL\REVISIONES.xlsx"
 
# source to destination
dest = shutil.copyfile(source, destination)
l_2G = 0
l_3G = 0
l_4G =0


driver = webdriver.Edge()
driver.get('https://gdcportal.net.nokia.com/home.php')
driver.maximize_window()

WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"//*[@class='loginSSO' and  @name='submit'  ]"))).click()
#time.sleep(20)
try:
    WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"//*[@id='menu_65_btn'  ]"))).click()
except:
    pass
time.sleep(3)
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"//*[@id='workorder' and  @class='menu-secundary-btn' ]"))).click()
time.sleep(6)
driver.find_element(By.XPATH,"//*[@name='word' and @id='word'  ]").send_keys(site)
driver.find_element(By.XPATH,"//*[@id='word']").send_keys(Keys.ENTER)

time.sleep(5)


try :
    
    path3=fnmatch.filter(os.listdir(str(path2)), '*-t735770-KPI_2G*.csv')
    path4=os.path.join(path1,str(site),str(path3[0]))

    GSM_raw_df = pd.read_csv(path4,sep=';',decimal=',')

    referencia = GSM_raw_df['PERIOD_START_TIME'].tail(1)

    all_data_2G_equal =  GSM_raw_df[GSM_raw_df['PERIOD_START_TIME'].values == referencia.values]
    all_data_2G = all_data_2G_equal[(all_data_2G_equal['TCH availability ratio'].str.contains('100.00'))==True]
    all_data_2G


    GSM_s_i=  all_data_2G.reset_index(drop=True)
    GSM_s_i['PERIOD_START_TIME'] = pd.to_datetime(GSM_s_i['PERIOD_START_TIME'])
    GSM_s_i['PERIOD_START_TIME'] = GSM_s_i['PERIOD_START_TIME'].dt.strftime('%d-%m-%Y')

    l_2G =len(GSM_s_i['PERIOD_START_TIME'])


    #####################################################################################################################

    net = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='60' ]")
    act  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='80']")

    nett =[]
    n= 0
    for k in  range(len(net)):  
        if net[k].text != str(site):
            nett.insert(n,net[k].text)
            n =n +1


    actt= [ ]
    n= 0
    coc = []                                                                                                                                                              
    for k in  range(len(act)):  
        if act[k].text  == 'INTEGRACION' or act[k].text == 'ANTENA' or act[k].text == 'INTEGRACION+e' or act[k].text=='ENCENDIDO' or act[k].text=='APAGADO'  or act[k].text=='ANTENA+RET' or act[k].text=='PRE-ENCENDIDO  ' : 
            actt.insert(k,act[k].text)
            coc.insert(k,act[k-1].text)
            n =n +1

    m=0        
    vector_coc = ['']*l_2G
    vector_catt = ['']*l_2G
    res = [x for x in range(len(nett)) if nett[x] == 'GSM' or nett[x] == 'GSM 900' ] 
    posiciones0 = GSM_s_i['BTS name'].index
    if posiciones0.size >0 and len(res) > 0:
        for n in  range(0,l_2G):
            if n in posiciones0:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =  actt[res[0]] 


    codigo = [site]*l_2G

    nombre  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='180']")
    uno = nombre[1].text
    tecnico =[uno.partition('\n')[0]]*l_2G

    REVISADO = ['DAVID ALVARADO']*l_2G


    GSM_s_i.insert(0, "COC", vector_coc, True) 
    GSM_s_i.insert(0, "ACT", vector_catt, True) 
    GSM_s_i.insert(0, "codigo", codigo, True) 
    GSM_s_i.insert(0, "tecnico", tecnico, True) 
    GSM_s_i.insert(0, "REVISADO", REVISADO, True)

    #####################################################################################################################3
    rev_path =  destination



    df_new = pd.DataFrame({'Col_C': GSM_s_i['ACT'], 
                           'Col_D': GSM_s_i['COC'],
                           'Col_E': GSM_s_i['codigo'],
                           'Col_J': GSM_s_i['tecnico'],
                           'Col_M': GSM_s_i['REVISADO'],
                           'Col_B': GSM_s_i['PERIOD_START_TIME'],'Col_F': GSM_s_i['BSC name'],
                           'Col_G': GSM_s_i['BCF name'],'Col_I': GSM_s_i['BTS name'], 
                           'Col_K': GSM_s_i['PERIOD_START_TIME'],

                           'Col_Q': GSM_s_i['TCH availability ratio'],'Col_R': GSM_s_i['TCH_CALL_REQ (C1026)'],
                           'Col_S': GSM_s_i['SDCCH success ratio'],'Col_T': GSM_s_i['TCH access'],
                           'Col_U': GSM_s_i['CSSR, voice'],'Col_V': GSM_s_i['CONVER_STARTED (c057015)'],
                           'Col_W': GSM_s_i['DROPPED_CALLS (c057007)'],'Col_X': GSM_s_i['TCH dropped conversation'],
                           'Col_Y': GSM_s_i['TCH traffic sum'],'Col_Z': GSM_s_i['UL GPRS RLC payload'],
                           'Col_AA': GSM_s_i['DL GPRS RLC payload'],'Col_AB': GSM_s_i['UL EGPRS RLC payload'],
                           'Col_AC': GSM_s_i['DL EGPRS RLC payload'],'Col_AD': GSM_s_i['TCH denied new call'],
                           'Col_AE': GSM_s_i['BSC_O_SUCC_HO (c004014)'],'Col_AF': GSM_s_i['BSC_I_SUCC_HO (c004010)']})

    wb = load_workbook(rev_path)

    ws = wb['Hoja1']

    for index, row in df_new.iterrows():
        cell = 'C%d'  % (index + 5 )
        ws[cell] = row[0]
    for index, row in df_new.iterrows():
        cell = 'D%d'  % (index + 5 )
        ws[cell] = row[1]
    for index, row in df_new.iterrows():
        cell = 'E%d'  % (index + 5 )
        ws[cell] = row[2]
    for index, row in df_new.iterrows():
        cell = 'J%d'  % (index + 5 )
        ws[cell] = row[3]
    for index, row in df_new.iterrows():
        cell = 'M%d'  % (index + 5 )
        ws[cell] = row[4]

    for index, row in df_new.iterrows():
        cell = 'B%d'  % (index + 5)
        ws[cell] = row[5]
    for index, row in df_new.iterrows():
        cell = 'F%d'  % (index + 5)
        ws[cell] = row[6]
    for index, row in df_new.iterrows():
        cell = 'G%d'  % (index + 5)
        ws[cell] = row[7]
    for index, row in df_new.iterrows():
        cell = 'I%d'  % (index + 5)
        ws[cell] = row[8]
    for index, row in df_new.iterrows():
        cell = 'K%d'  % (index + 5)
        ws[cell] = row[9]
    for index, row in df_new.iterrows():
        cell = 'Q%d'  % (index + 5)
        ws[cell] = row[10]
    for index, row in df_new.iterrows():
        cell = 'R%d'  % (index + 5)
        ws[cell] = row[11]
    for index, row in df_new.iterrows():
        cell = 'S%d'  % (index + 5)
        ws[cell] = row[12]    
    for index, row in df_new.iterrows():
        cell = 'T%d'  % (index + 5)
        ws[cell] = row[13]    
    for index, row in df_new.iterrows():
        cell = 'U%d'  % (index + 5)
        ws[cell] = row[14]    
    for index, row in df_new.iterrows():
        cell = 'V%d'  % (index + 5)
        ws[cell] = row[15]    
    for index, row in df_new.iterrows():
        cell = 'W%d'  % (index + 5)
        ws[cell] = row[16]    
    for index, row in df_new.iterrows():
        cell = 'X%d'  % (index + 5)
        ws[cell] = row[17]    
    for index, row in df_new.iterrows():
        cell = 'Y%d'  % (index + 5)
        ws[cell] = row[18] 
    for index, row in df_new.iterrows():
        cell = 'Z%d'  % (index + 5)
        ws[cell] = row[19]    
    for index, row in df_new.iterrows():
        cell = 'AA%d'  % (index + 5)
        ws[cell] = row[20]    
    for index, row in df_new.iterrows():
        cell = 'AB%d'  % (index + 5)
        ws[cell] = row[21]    
    for index, row in df_new.iterrows():
        cell = 'AC%d'  % (index + 5)
        ws[cell] = row[22]
    for index, row in df_new.iterrows():
        cell = 'AD%d'  % (index + 5)
        ws[cell] = row[23]

    for index, row in df_new.iterrows():
        cell = 'AE%d'  % (index + 5)
        ws[cell] = row[24]

    for index, row in df_new.iterrows():
        cell = 'AF%d'  % (index + 5)

        ws[cell] = row[25]

    wb.save(rev_path)    

    wb.save(rev_path)
    
    #all_data['TCH availability ratio'].values == 100
#all_data['TCH_CALL_REQ (C1026)'].values >  0
#all_data['SDCCH success ratio'].values > '80.00'
#all_data['TCH access'].values > '90.00'
#all_data['CSSR, voice'].values == '80.00'
##all_data['TCH dropped conversation'].values < '10.00'
#all_data['TCH traffic sum'].values > '0.00'
#all_data['UL GPRS RLC payload'].values > '0.00'
#all_data['DL GPRS RLC payload'].values > '0.00'
#all_data['UL EGPRS RLC payload'].values > '0.00'
#all_data['DL EGPRS RLC payload'].values > '0.00'
#all_data['TCH denied new call'].values < '10.00'

except:
    print(' no hay GSM')
  #####################################################################################################################
 #####################################################################################################################
     #####################################################################################################################
         #####################################################################################################################
             #####################################################################################################################
                 #####################################################################################################################
                    
    
    
try :   
    path3=fnmatch.filter(os.listdir(str(path2)), '*-t735770-KPI_3G*.csv')
    path4=os.path.join(path1,str(site),str(path3[0]))

    UMTS_raw_df = pd.read_csv(path4,sep=';',decimal=',')

    referencia = UMTS_raw_df['PERIOD_START_TIME'].tail(1)

    all_data_3G_equal = UMTS_raw_df[UMTS_raw_df['PERIOD_START_TIME'].values == referencia.values]
    all_data_3G = all_data_3G_equal[(all_data_3G_equal['Cell Availability'].str.contains('100.00'))==True]


    UMTS_s_i=  all_data_3G.reset_index(drop=True)
    UMTS_s_i['PERIOD_START_TIME'] = pd.to_datetime(UMTS_s_i['PERIOD_START_TIME'])
    UMTS_s_i['PERIOD_START_TIME'] = UMTS_s_i['PERIOD_START_TIME'].dt.strftime('%d-%m-%Y')

    UMTS_s_i
    l_3G = len(UMTS_s_i[(UMTS_s_i['Cell Availability'].str.contains('100.00'))==True])


    #####################################################################################################################

    net = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='60' ]")
    act  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='80']")

    nett =[]
    n= 0
    for k in  range(len(net)):  
        if net[k].text != str(site):
            nett.insert(n,net[k].text)
            n =n +1


    actt= [ ]
    n= 0
    coc = []
    for k in  range(len(act)):  
        if act[k].text  == 'INTEGRACION' or act[k].text == 'ANTENA' or act[k].text == 'INTEGRACION+e' or act[k].text=='ENCENDIDO' or act[k].text=='APAGADO'  or act[k].text=='ANTENA+RET' or act[k].text=='PRE-ENCENDIDO  ': 
            actt.insert(k,act[k].text)
            coc.insert(k,act[k-1].text)
            n =n +1

    m=0        
    vector_coc = ['']*l_3G
    vector_catt = ['']*l_3G
    posicionesp9 = []

    res = [x for x in range(len(nett)) if nett[x] == 'UMTS 900' or nett[x] == 'U900' or nett[x] == 'UMTS900' ]
    posicionesp9 = UMTS_s_i[UMTS_s_i['WCEL name'].str.contains('p9')== True].index
    if posicionesp9.size > 0 and len(res) > 0:
        for n in  range(0,l_3G):
            if n in posicionesp9:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =  actt[res[0]] 

    posicionesp8 = []           
    res = [x for x in range(len(nett)) if nett[x] == 'UMTS 900' or nett[x] == 'U900' or nett[x] == 'UMTS900' ]
    posicionesp8 = UMTS_s_i[UMTS_s_i['WCEL name'].str.contains('p8')== True].index
    if posicionesp8.size > 0 and len(res) > 0:
        for n in  range(0,l_3G):
            if n in posiciones8:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =  actt[res[0]]             


    posicionesp1 = []          
    res = [x for x in range(len(nett)) if nett[x] == 'UMTS' or nett[x] == 'UMTS 2100' or nett[x] == 'U2100' or nett[x] == 'UMTS2100' ]
    posicionesp1 = UMTS_s_i[UMTS_s_i['WCEL name'].str.contains('p1')== True].index
    if posicionesp1.size > 0  and len(res) > 0 :
        for n in  range(0,l_3G):
            if n in posicionesp1:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =  actt[res[0]] 

    posicionesp2 = []             
    res = [x for x in range(len(nett)) if  nett[x] == 'UMTS' or nett[x] == 'UMTS 2100' or nett[x] == 'U2100' or nett[x] == 'UMTS2100' ]
    posicionesp2 = UMTS_s_i[UMTS_s_i['WCEL name'].str.contains('p2')== True].index
    if posicionesp2.size > 0 and len(res) > 0:
        for n in  range(0,l_3G):
            if n in posicionesp2:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =  actt[res[0]]  



    codigo = [site]*l_3G

    nombre  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='180']")
    uno = nombre[1].text
    tecnico =[uno.partition('\n')[0]]*l_3G

    REVISADO = ['DAVID ALVARADO']*l_3G


    UMTS_s_i.insert(0, "COC", vector_coc, True) 
    UMTS_s_i.insert(0, "ACT", vector_catt, True) 
    UMTS_s_i.insert(0, "codigo", codigo, True)  
    UMTS_s_i.insert(0, "tecnico", tecnico, True) 
    UMTS_s_i.insert(0, "REVISADO", REVISADO, True)
    UMTS_s_i = UMTS_s_i.sort_values(by='WCEL name').reset_index(drop=True)

    #####################################################################################################################3
    rev_path =  destination


    df_new = pd.DataFrame({'Col_B': UMTS_s_i['PERIOD_START_TIME'],'Col_F': UMTS_s_i['RNC name'],
                           'Col_G': UMTS_s_i['WBTS name'],'Col_H': UMTS_s_i['WBTS ID'], 
                           'Col_I': UMTS_s_i['WCEL name'],'Col_K': UMTS_s_i['PERIOD_START_TIME'],

                           'Col_AG': UMTS_s_i['Cell Availability'],'Col_AH': UMTS_s_i['Establecimientos voz_NIR118'],
                           'Col_AI': UMTS_s_i['Establecimientos_PS_NIR118'],'Col_AJ': UMTS_s_i['Tasa de fallo en el establecimiento de conexiones RRC (CS)_NIR118'],
                           'Col_AK': UMTS_s_i['Tasa de fallo en el establecimiento de conexiones RRC (PS)_NIR118'],'Col_AL': UMTS_s_i['Tasa de conexiones  caidas (voz)_NIR118'],
                           'Col_AM': UMTS_s_i['Tasa de conexiones  caidas PS_NIR118'],'Col_AN': UMTS_s_i['Tasa de fallo en el establecimiento RAB Voz_NIR118'],
                           'Col_AO': UMTS_s_i['Total CS traffic - Erl'],'Col_AP': UMTS_s_i['Fallos establecimiento HSDPA_NIR118'],
                           'Col_AQ': UMTS_s_i['Fallos establecimiento HSUPA_NIR118'],
                           'Col_C': UMTS_s_i['ACT'], 
                           'Col_D': UMTS_s_i['COC'],
                           'Col_E': UMTS_s_i['codigo'],
                           'Col_J': UMTS_s_i['tecnico'],
                           'Col_M': UMTS_s_i['REVISADO']

                          })




    wb = load_workbook(rev_path)

    ws = wb['Hoja1']

    for index, row in df_new.iterrows():
        cell = 'B%d'  % (index + 5+l_2G)
        ws[cell] = row[0]
    for index, row in df_new.iterrows():
        cell = 'F%d'  % (index + 5+l_2G)
        ws[cell] = row[1]
    for index, row in df_new.iterrows():
        cell = 'G%d'  % (index + 5+l_2G)
        ws[cell] = row[2]
    for index, row in df_new.iterrows():
        cell = 'H%d'  % (index + 5+l_2G)
        ws[cell] = row[3]
    for index, row in df_new.iterrows():
        cell = 'I%d'  % (index + 5+l_2G)
        ws[cell] = row[4]
    for index, row in df_new.iterrows():
        cell = 'K%d'  % (index + 5+l_2G)
        ws[cell] = row[5]

    for index, row in df_new.iterrows():
        cell = 'AG%d'  % (index + 5+l_2G)
        ws[cell] = row[6]
    for index, row in df_new.iterrows():
        cell = 'AH%d'  % (index + 5+l_2G)
        ws[cell] = row[7]
    for index, row in df_new.iterrows():
        cell = 'AI%d'  % (index + 5+l_2G)
        ws[cell] = row[8]
    for index, row in df_new.iterrows():
        cell = 'AJ%d'  % (index + 5+l_2G)
        ws[cell] = row[9]
    for index, row in df_new.iterrows():
        cell = 'AK%d'  % (index + 5+l_2G)
        ws[cell] = row[10]
    for index, row in df_new.iterrows():
        cell = 'AL%d'  % (index + 5+l_2G)
        ws[cell] = row[11]
    for index, row in df_new.iterrows():
        cell = 'AM%d'  % (index + 5+l_2G)
        ws[cell] = row[12]
    for index, row in df_new.iterrows():
        cell = 'AN%d'  % (index + 5+l_2G)
        ws[cell] = row[13]
    for index, row in df_new.iterrows():
        cell = 'AO%d'  % (index + 5+l_2G)
        ws[cell] = row[14]
    for index, row in df_new.iterrows():
        cell = 'AP%d'  % (index + 5+l_2G)
        ws[cell] = row[15]

    for index, row in df_new.iterrows():
        cell = 'AQ%d'  % (index + 5+l_2G)
        ws[cell] = row[16]
    for index, row in df_new.iterrows():
        cell = 'C%d'  % (index + 5+l_2G)
        ws[cell] = row[17]
    for index, row in df_new.iterrows():
        cell = 'D%d'  % (index + 5+l_2G)
        ws[cell] = row[18]
    for index, row in df_new.iterrows():
        cell = 'E%d'  % (index + 5+l_2G)
        ws[cell] = row[19]
    for index, row in df_new.iterrows():
        cell = 'J%d'  % (index + 5+l_2G)
        ws[cell] = row[20]
    for index, row in df_new.iterrows():
        cell = 'M%d'  % (index + 5+l_2G)
        ws[cell] = row[21]

    wb.save(rev_path)
    
except:
    print(' no hay UMTS')    
  #####################################################################################################################
 #####################################################################################################################
     #####################################################################################################################
         #####################################################################################################################
             #####################################################################################################################
                 #####################################################################################################################

try:
    
    path3=fnmatch.filter(os.listdir(str(path2)), '*-t735770-KPI_4G*.csv')
    path4=os.path.join(path1,str(site),str(path3[0]))

    LTE_raw_df = pd.read_csv(path4,sep=';',decimal=',')

    referencia = LTE_raw_df['PERIOD_START_TIME'].tail(1)

    all_data_4G_equal = LTE_raw_df[LTE_raw_df['PERIOD_START_TIME'].values == referencia.values]
    all_data_4G = all_data_4G_equal[(all_data_4G_equal['Cell Avail'].str.contains('100.00'))==True]


    LTE_s_i=  all_data_4G.reset_index(drop=True)
    LTE_s_i['PERIOD_START_TIME'] = pd.to_datetime(LTE_s_i['PERIOD_START_TIME'])
    LTE_s_i['PERIOD_START_TIME'] = LTE_s_i['PERIOD_START_TIME'].dt.strftime('%d-%m-%Y')

    l_4G = len(LTE_s_i[(LTE_s_i['Cell Avail'].str.contains('100.00'))==True])


    #####################################################################################################################

    net = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='60' ]")
    act  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='80']")

    nett =[]
    n= 0
    for k in  range(len(net)):  
        if net[k].text != str(site):
            nett.insert(n,net[k].text)
            n =n +1

    net = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='60' ]")
    act  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='80']")

    actt= []
    n= 0
    coc = []
    for k in  range(len(act)):                                                                                                                                                                         
        if act[k].text  == 'INTEGRACION' or act[k].text == 'ANTENA' or act[k].text == 'INTEGRACION+e' or act[k].text=='ENCENDIDO' or act[k].text=='APAGADO'  or act[k].text=='ANTENA+RET' or act[k].text=='PRE-ENCENDIDO  ' : 
            actt.insert(k,act[k].text)
            coc.insert(k,act[k-1].text)
            n =n +1

    m=0        
    vector_coc = ['']*l_4G
    vector_catt = ['']*l_4G



    res = [x for x in range(len(nett)) if nett[x] == 'LTE 1800'or nett[x] == 'LTE1800']
    posiciones0 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]0')== True].index
    if posiciones0.size > 0 and len(res) > 0:
        for n in  range(0,l_4G):
            if n in posiciones0:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =  actt[res[0]] 



    res = [x for x in range(len(nett)) if nett[x] == 'LTE 800']
    posiciones1 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]1')== True].index
    if posiciones1.size > 0  and len(res) > 0: 
        for n in  range(0,l_4G):
            if n in posiciones1:
                if res != []:
                    vector_coc[n]= coc[res[0]] 
                    vector_catt[n] = actt[res[0]] 


    res = [x for x in range(len(nett)) if nett[x] == 'LTE 2600' or nett[x] == 'LTE2600' ]
    posiciones2 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]2')== True].index
    if posiciones2.size > 0  and len(res) > 0:  
        for n in  range(0,l_4G):
            if n in posiciones2:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =actt[res[0]] 



    res = [x for x in range(len(nett)) if nett[x] == 'LTE 900' or nett[x] == 'LTE900' ]
    posiciones3 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]3')== True].index
    if posiciones3.size > 0  and len(res) > 0:   
        for n in  range(0,l_4G):
            if n in posiciones3:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =actt[res[0]] 

    res = [x for x in range(len(nett)) if nett[x] == 'LTE 2100' or nett[x] == 'LTE2100' ]
    posiciones4 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]4')== True].index
    if posiciones4.size > 0  and len(res) > 0:   
        for n in  range(0,l_4G):
            if n in posiciones4:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =actt[res[0]] 

    res = [x for x in range(len(nett)) if nett[x] == 'LTE 700' or nett[x] == 'LTE700' ]
    posiciones5 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]5')== True].index
    if posiciones5.size > 0  and len(res) > 0:   
        for n in  range(0,l_4G):
            if n in posiciones5:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =actt[res[0]] 

    res = [x for x in range(len(nett)) if nett[x] == 'LTE 700' or nett[x] == 'LTE700' ]
    posiciones7 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]7')== True].index
    if posiciones7.size > 0  and len(res) > 0:   
        for n in  range(0,l_4G):
            if n in posiciones7:  
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] =actt[res[0]] 



    codigo = [site]*l_4G

    nombre  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='180']")
    uno = nombre[1].text
    tecnico =[uno.partition('\n')[0]]*l_4G

    REVISADO = ['DAVID ALVARADO']*l_4G

    provincia  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='50']")
    provnica_t = provincia[1].text
    privincia_lte = [provnica_t]*l_4G


    LTE_s_i.insert(0, "COC", vector_coc, True) 
    LTE_s_i.insert(0, "ACT", vector_catt, True) 
    LTE_s_i.insert(0, "codigo", codigo, True) 
    LTE_s_i.insert(0, "privincia_lte", privincia_lte, True) 
    LTE_s_i.insert(0, "tecnico", tecnico, True) 
    LTE_s_i.insert(0, "REVISADO", REVISADO, True)

    #####################################################################################################################3
    rev_path =  destination



    df_new = pd.DataFrame({'Col_B': LTE_s_i['PERIOD_START_TIME'], 
                           'Col_G': LTE_s_i['LNBTS name'],'Col_I': LTE_s_i['LNCEL name'], 
                           'Col_K': LTE_s_i['PERIOD_START_TIME'],
                           'Col_AR': LTE_s_i['Cell Avail'],'Col_AS': LTE_s_i['RRC stp att'],
                           'Col_AT': LTE_s_i['Total E-UTRAN RRC conn stp SR'],'Col_AU': LTE_s_i['E-UTRAN E-RAB Setup Attempts'],
                           'Col_AV': LTE_s_i['E-UTRAN E-RAB stp SR'],'Col_AW': LTE_s_i['E-RAB DR, RAN view'],
                           'Col_AX': LTE_s_i['CSFB att UE iddle mode'],'Col_AY': LTE_s_i['CSFB att UE conn mode'],
                           'Col_AZ': LTE_s_i['PDCP_SDU_VOL_DL (M8012C20)'],
                           'Col_C': LTE_s_i['ACT'], 
                           'Col_D': LTE_s_i['COC'],
                           'Col_E': LTE_s_i['codigo'],
                           'Col_F': LTE_s_i['privincia_lte'],
                           'Col_J': LTE_s_i['tecnico'],
                           'Col_M': LTE_s_i['REVISADO']


                          })


    wb = load_workbook(rev_path)

    ws = wb['Hoja1']


    for index, row in df_new.iterrows():
        cell = 'B%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[0]
    for index, row in df_new.iterrows():
        cell = 'G%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[1]
    for index, row in df_new.iterrows():
        cell = 'I%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[2]
    for index, row in df_new.iterrows():
        cell = 'K%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[3]
    for index, row in df_new.iterrows():
        cell = 'AR%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[4]
    for index, row in df_new.iterrows():
        cell = 'AS%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[5]
    for index, row in df_new.iterrows():
        cell = 'AT%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[6]
    for index, row in df_new.iterrows():
        cell = 'AU%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[7]
    for index, row in df_new.iterrows():
        cell = 'AV%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[8]
    for index, row in df_new.iterrows():
        cell = 'AW%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[9]
    for index, row in df_new.iterrows():
        cell = 'AX%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[10]
    for index, row in df_new.iterrows():
        cell = 'AY%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[11]
    for index, row in df_new.iterrows():
        cell = 'AZ%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[12]
    for index, row in df_new.iterrows():
        cell = 'C%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[13]
    for index, row in df_new.iterrows():
        cell = 'D%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[14]
    for index, row in df_new.iterrows():
        cell = 'E%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[15]
    for index, row in df_new.iterrows():
        cell = 'F%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[16]
    for index, row in df_new.iterrows():
        cell = 'J%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[17]
    for index, row in df_new.iterrows():
        cell = 'M%d'  % (index + 5+l_2G+l_3G)
        ws[cell] = row[18]


    wb.save(rev_path)

except:
    print(' no hay 4G')  
    
driver.close()
print('terminado el proceso de KPISs')


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[8]:


site = 700025
# Destination path
destination = r'C:\Users\davalvar\OneDrive - Nokia\Snapshots\SNAPSHOTS 2022\10 Outubro\13\0700025\0700025_KPIS.xlsx'

#  r"C:\Users\davalvar\OneDrive - Nokia\Snapshots\SNAPSHOTS 2022\10 Outubro\04\3600112

path1 = r'C:\Users\davalvar\Desktop'
path2=os.path.join(path1,str(site))
# Source path
source = r"C:\Users\davalvar\Desktop\Explicaciones Nokia\EXCEL\REVISIONES.xlsx"
 
# source to destination
dest = shutil.copyfile(source, destination)
l_2G = 0
l_3G = 0
l_4G =0


driver = webdriver.Edge()
driver.get('https://gdcportal.net.nokia.com/home.php')
driver.maximize_window()
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"//*[@class='loginSSO' and  @name='submit'  ]"))).click()
time.sleep(3)
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"//*[@id='menu_65_btn'  ]"))).click()
WebDriverWait(driver, 5).until(EC.element_to_be_clickable((By.XPATH,"//*[@id='workorder' and  @class='menu-secundary-btn' ]"))).click()
time.sleep(6)
driver.find_element(By.XPATH,"//*[@name='word' and @id='word'  ]").send_keys(site)
driver.find_element(By.XPATH,"//*[@id='word']").send_keys(Keys.ENTER)

time.sleep(5)


# In[22]:


path3=fnmatch.filter(os.listdir(str(path2)), '*-t735770-KPI_4G*.csv')
path4=os.path.join(path1,str(site),str(path3[0]))

LTE_raw_df = pd.read_csv(path4,sep=';',decimal=',')

referencia = LTE_raw_df['PERIOD_START_TIME'].tail(1)

all_data_4G_equal = LTE_raw_df[LTE_raw_df['PERIOD_START_TIME'].values == referencia.values]
all_data_4G = all_data_4G_equal[(all_data_4G_equal['Cell Avail'].str.contains('100.00'))==True]


LTE_s_i=  all_data_4G.reset_index(drop=True)
LTE_s_i['PERIOD_START_TIME'] = pd.to_datetime(LTE_s_i['PERIOD_START_TIME'])
LTE_s_i['PERIOD_START_TIME'] = LTE_s_i['PERIOD_START_TIME'].dt.strftime('%d-%m-%Y')

l_4G = len(LTE_s_i[(LTE_s_i['Cell Avail'].str.contains('100.00'))==True])


#####################################################################################################################

net = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='60' ]")
act  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='80']")

nett =[]
n= 0
for k in  range(len(net)):  
    if net[k].text != str(site):
        nett.insert(n,net[k].text)
        n =n +1

net = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='60' ]")
act  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='80']")

actt= []
n= 0
coc = []
for k in  range(len(act)):                                                                                                                                                                         
    if act[k].text  == 'INTEGRACION' or act[k].text == 'ANTENA' or act[k].text == 'INTEGRACION+e' or act[k].text=='ENCENDIDO' or act[k].text=='APAGADO'  or act[k].text=='ANTENA+RET' or act[k].text=='PRE-ENCENDIDO  ' : 
        actt.insert(k,act[k].text)
        coc.insert(k,act[k-1].text)
        n =n +1

m=0        
vector_coc = ['']*l_4G
vector_catt = ['']*l_4G



res = [x for x in range(len(nett)) if nett[x] == 'LTE 1800'or nett[x] == 'LTE1800']
posiciones0 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]0')== True].index
if posiciones0.size > 0 and len(res) > 0:
    for n in  range(0,l_4G):
        if n in posiciones0:  
            vector_coc[n]= coc[res[0]] 
            vector_catt[n] =  actt[res[0]] 



res = [x for x in range(len(nett)) if nett[x] == 'LTE 800']
posiciones1 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]1')== True].index
if posiciones1.size > 0  and len(res) > 0: 
    for n in  range(0,l_4G):
        if n in posiciones1:
            if res != []:
                vector_coc[n]= coc[res[0]] 
                vector_catt[n] = actt[res[0]] 


res = [x for x in range(len(nett)) if nett[x] == 'LTE 2600' or nett[x] == 'LTE2600' ]
posiciones2 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]2')== True].index
if posiciones2.size > 0  and len(res) > 0:  
    for n in  range(0,l_4G):
        if n in posiciones2:  
            vector_coc[n]= coc[res[0]] 
            vector_catt[n] =actt[res[0]] 



res = [x for x in range(len(nett)) if nett[x] == 'LTE 900' or nett[x] == 'LTE900' ]
posiciones3 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]3')== True].index
if posiciones3.size > 0  and len(res) > 0:   
    for n in  range(0,l_4G):
        if n in posiciones3:  
            vector_coc[n]= coc[res[0]] 
            vector_catt[n] =actt[res[0]] 

res = [x for x in range(len(nett)) if nett[x] == 'LTE 2100' or nett[x] == 'LTE2100' ]
posiciones4 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]4')== True].index
if posiciones4.size > 0  and len(res) > 0:   
    for n in  range(0,l_4G):
        if n in posiciones4:  
            vector_coc[n]= coc[res[0]] 
            vector_catt[n] =actt[res[0]] 

res = [x for x in range(len(nett)) if nett[x] == 'LTE 700' or nett[x] == 'LTE700' ]
posiciones5 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]5')== True].index
if posiciones5.size > 0  and len(res) > 0:   
    for n in  range(0,l_4G):
        if n in posiciones5:  
            vector_coc[n]= coc[res[0]] 
            vector_catt[n] =actt[res[0]] 

res = [x for x in range(len(nett)) if nett[x] == 'LTE 700' or nett[x] == 'LTE700' ]
posiciones7 = LTE_s_i[LTE_s_i['LNCEL name'].str.contains('_0[0-9]7')== True].index
if posiciones7.size > 0  and len(res) > 0:   
    for n in  range(0,l_4G):
        if n in posiciones7:  
            vector_coc[n]= coc[res[0]] 
            vector_catt[n] =actt[res[0]] 



codigo = [site]*l_4G

nombre  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='180']")
uno = nombre[0].text
tecnico =[uno.partition('\n')[0]]*l_4G

REVISADO = ['DAVID ALVARADO']*l_4G

provincia  = driver.find_elements(By.XPATH,"//*[@class='gridColumm' and @width='50']")
provnica_t = provincia[1].text
privincia_lte = [provnica_t]*l_4G


LTE_s_i.insert(0, "COC", vector_coc, True) 
LTE_s_i.insert(0, "ACT", vector_catt, True) 
LTE_s_i.insert(0, "codigo", codigo, True) 
LTE_s_i.insert(0, "privincia_lte", privincia_lte, True) 
LTE_s_i.insert(0, "tecnico", tecnico, True) 
LTE_s_i.insert(0, "REVISADO", REVISADO, True)

#####################################################################################################################3
rev_path =  destination



df_new = pd.DataFrame({'Col_B': LTE_s_i['PERIOD_START_TIME'], 
                       'Col_G': LTE_s_i['LNBTS name'],'Col_I': LTE_s_i['LNCEL name'], 
                       'Col_K': LTE_s_i['PERIOD_START_TIME'],
                       'Col_AR': LTE_s_i['Cell Avail'],'Col_AS': LTE_s_i['RRC stp att'],
                       'Col_AT': LTE_s_i['Total E-UTRAN RRC conn stp SR'],'Col_AU': LTE_s_i['E-UTRAN E-RAB Setup Attempts'],
                       'Col_AV': LTE_s_i['E-UTRAN E-RAB stp SR'],'Col_AW': LTE_s_i['E-RAB DR, RAN view'],
                       'Col_AX': LTE_s_i['CSFB att UE iddle mode'],'Col_AY': LTE_s_i['CSFB att UE conn mode'],
                       'Col_AZ': LTE_s_i['PDCP_SDU_VOL_DL (M8012C20)'],
                       'Col_C': LTE_s_i['ACT'], 
                       'Col_D': LTE_s_i['COC'],
                       'Col_E': LTE_s_i['codigo'],
                       'Col_F': LTE_s_i['privincia_lte'],
                       'Col_J': LTE_s_i['tecnico'],
                       'Col_M': LTE_s_i['REVISADO']


                      })


wb = load_workbook(rev_path)

ws = wb['Hoja1']


for index, row in df_new.iterrows():
    cell = 'B%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[0]
for index, row in df_new.iterrows():
    cell = 'G%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[1]
for index, row in df_new.iterrows():
    cell = 'I%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[2]
for index, row in df_new.iterrows():
    cell = 'K%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[3]
for index, row in df_new.iterrows():
    cell = 'AR%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[4]
for index, row in df_new.iterrows():
    cell = 'AS%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[5]
for index, row in df_new.iterrows():
    cell = 'AT%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[6]
for index, row in df_new.iterrows():
    cell = 'AU%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[7]
for index, row in df_new.iterrows():
    cell = 'AV%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[8]
for index, row in df_new.iterrows():
    cell = 'AW%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[9]
for index, row in df_new.iterrows():
    cell = 'AX%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[10]
for index, row in df_new.iterrows():
    cell = 'AY%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[11]
for index, row in df_new.iterrows():
    cell = 'AZ%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[12]
for index, row in df_new.iterrows():
    cell = 'C%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[13]
for index, row in df_new.iterrows():
    cell = 'D%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[14]
for index, row in df_new.iterrows():
    cell = 'E%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[15]
for index, row in df_new.iterrows():
    cell = 'F%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[16]
for index, row in df_new.iterrows():
    cell = 'J%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[17]
for index, row in df_new.iterrows():
    cell = 'M%d'  % (index + 5+l_2G+l_3G)
    ws[cell] = row[18]


wb.save(rev_path)

def fecha():
    import time
    fecha = time.strftime("%d/%m/%y")
    dia = fecha[0:2]
    mes_aux = fecha[3:5]
    month = ''
    if  mes_aux == '01':
        month ='01 Janeiro'
    elif mes_aux == '02':
        month = '02 Fevereiro'
    elif mes_aux == '3':
        month = '03 Março'
    elif mes_aux == '04':
        month= '04 Abril'
    elif mes_aux == '05':
        month= '05 Maio'
    elif mes_aux == '06':
        month = '06 Junho'
    elif mes_aux == '07':
        month = '07 Julho'
    elif mes_aux == '08':
        month = '08 Agosto'
    elif mes_aux == '09':
        month= '09 Setembro'
    elif mes_aux == '10':
        month= '10 Outubro'
    elif mes_aux == '11':
        month= '11 Novembro'
    elif mes_aux == '12':
        month= '12 Dezembro'
        
    return(dia,month)


# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[ ]:





# In[47]:




